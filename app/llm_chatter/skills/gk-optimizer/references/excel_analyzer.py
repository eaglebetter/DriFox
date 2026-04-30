# -*- coding: utf-8 -*-
"""
Excel结构分析脚本 - 自动分析工况Excel文件结构

功能：
1. 分析Excel文件的Sheet结构、列名、列索引
2. 定位"寻优上限"、"寻优下限"列的位置
3. 识别工况参数列范围
4. 输出结构化分析报告，供生成专家规则脚本使用

使用方法：
    python excel_analyzer.py <excel文件路径>
    python excel_analyzer.py                                    # 扫描当前目录
    python excel_analyzer.py <excel文件路径> --json             # 输出JSON格式
    python excel_analyzer.py <excel文件路径> --sample 3         # 每行显示3个样例值
"""

import openpyxl
import os
import sys
import json
import argparse
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict


@dataclass
class ColumnInfo:
    """列信息"""
    index: int  # 0-based
    name: str
    sample_values: List[Any]
    is_upper_limit: bool = False
    is_lower_limit: bool = False


@dataclass
class SheetAnalysis:
    """Sheet分析结果"""
    name: str
    rows: int
    cols: int
    columns: List[ColumnInfo]
    upper_limit_col_idx: Optional[int] = None
    lower_limit_col_idx: Optional[int] = None
    condition_param_range: tuple = None  # (start_col, end_col) 0-based


class ExcelAnalyzer:
    """Excel结构分析器"""
    
    def __init__(self, sample_rows: int = 3):
        """
        初始化分析器
        
        Args:
            sample_rows: 每个列采集的样例值数量
        """
        self.sample_rows = sample_rows
    
    def analyze_file(self, filepath: str) -> Dict[str, Any]:
        """
        分析Excel文件
        
        Args:
            filepath: Excel文件路径
            
        Returns:
            分析结果字典
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        result = {
            "file": filepath,
            "filename": os.path.basename(filepath),
            "sheets": []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_analysis = self._analyze_sheet(ws)
            sheet_analysis.name = sheet_name
            result["sheets"].append(asdict(sheet_analysis))
        
        return result
    
    def _analyze_sheet(self, ws) -> SheetAnalysis:
        """分析单个Sheet"""
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        
        analysis = SheetAnalysis(
            name="",
            rows=rows,
            cols=cols,
            columns=[]
        )
        
        if rows == 0 or cols == 0:
            return analysis
        
        # 读取表头行
        headers = []
        for col in range(1, cols + 1):
            headers.append(ws.cell(row=1, column=col).value)
        
        # 读取样例数据行（最多5行）
        sample_data_rows = []
        for row in range(2, min(rows + 1, 7)):
            row_data = [ws.cell(row=row, column=col).value for col in range(1, cols + 1)]
            sample_data_rows.append(row_data)
        
        # 分析每一列
        upper_limit_col_idx = None
        lower_limit_col_idx = None
        
        for col_idx, header in enumerate(headers):
            col_name = str(header) if header is not None else f"列{col_idx+1}"
            
            # 采集样例值
            sample_values = []
            for row_data in sample_data_rows:
                if col_idx < len(row_data) and row_data[col_idx] is not None:
                    val = row_data[col_idx]
                    # 限制样例值长度
                    if isinstance(val, str) and len(val) > 50:
                        val = val[:50] + "..."
                    sample_values.append(val)
            
            col_info = ColumnInfo(
                index=col_idx,
                name=col_name,
                sample_values=sample_values[:self.sample_rows]
            )
            
            # 检查是否是上限/下限列
            if "上限" in col_name and "寻优" in col_name:
                col_info.is_upper_limit = True
                upper_limit_col_idx = col_idx
            elif "下限" in col_name and "寻优" in col_name:
                col_info.is_lower_limit = True
                lower_limit_col_idx = col_idx
            
            analysis.columns.append(col_info)
        
        analysis.upper_limit_col_idx = upper_limit_col_idx
        analysis.lower_limit_col_idx = lower_limit_col_idx
        
        # 计算工况参数范围（从第2列开始到下限列之前）
        if lower_limit_col_idx is not None:
            analysis.condition_param_range = (1, lower_limit_col_idx)  # 0-based
        else:
            analysis.condition_param_range = (1, cols)
        
        return analysis
    
    def generate_report(self, analysis: Dict) -> str:
        """
        生成文本格式的分析报告
        
        Args:
            analysis: 分析结果
            
        Returns:
            格式化的报告文本
        """
        lines = []
        lines.append("=" * 70)
        lines.append(f"Excel结构分析报告: {analysis['filename']}")
        lines.append("=" * 70)
        lines.append("")
        
        for sheet in analysis["sheets"]:
            lines.append(f"【Sheet: {sheet['name']}】")
            lines.append(f"  行数: {sheet['rows']}, 列数: {sheet['cols']}")
            
            if sheet['upper_limit_col_idx'] is not None:
                lines.append(f"  ★ 寻优上限列: 第{sheet['upper_limit_col_idx']+1}列")
            if sheet['lower_limit_col_idx'] is not None:
                lines.append(f"  ★ 寻优下限列: 第{sheet['lower_limit_col_idx']+1}列")
            
            if sheet['condition_param_range']:
                start, end = sheet['condition_param_range']
                lines.append(f"  ○ 工况参数范围: 第{start+1}列 ~ 第{end}列 (共{end-start}列)")
            
            lines.append("")
            lines.append("  列结构:")
            lines.append("  " + "-" * 60)
            lines.append(f"  {'序号':<4} {'列名':<30} {'样例值'}")
            lines.append("  " + "-" * 60)
            
            for col in sheet['columns']:
                col_marker = ""
                if col['is_upper_limit']:
                    col_marker = "[上限]"
                elif col['is_lower_limit']:
                    col_marker = "[下限]"
                
                sample_str = ", ".join([str(v) for v in col['sample_values'][:2]])
                if len(sample_str) > 25:
                    sample_str = sample_str[:25] + "..."
                
                lines.append(
                    f"  {col['index']+1:<4} {col['name']:<45} {col_marker}{sample_str}"
                )
            
            lines.append("")
        
        lines.append("=" * 70)
        lines.append("【脚本生成建议】")
        lines.append("=" * 70)
        
        # 生成标签映射建议
        lines.append("\n标签映射表建议:")
        lines.append("LABEL_MAPPING = {")
        
        for sheet in analysis["sheets"]:
            # 查找非标准列（可能是测点相关的列）
            for col in sheet["columns"]:
                if col["index"] > 0:  # 跳过第一列
                    col_name = col["name"]
                    # 检查是否是数值类型的列
                    has_numeric = False
                    for val in col["sample_values"]:
                        if isinstance(val, (int, float)):
                            has_numeric = True
                            break
                    
                    if has_numeric and "寻优" not in col_name and "原始" not in col_name:
                        lines.append(f'    "{col_name}": "简化标签",')
        
        lines.append("}")
        
        # 生成列索引建议
        lines.append("\n列索引定义建议 (0-based):")
        lines.append("# 工况参数列索引范围")
        
        for sheet in analysis["sheets"]:
            if sheet["condition_param_range"]:
                start, end = sheet["condition_param_range"]
                lines.append(f"# {sheet['name']}: 第{start+1}列 ~ 第{end}列")
        
        return "\n".join(lines)


def find_excel_files(path: Optional[str] = None) -> List[str]:
    """查找Excel文件"""
    if path and os.path.isfile(path):
        return [path]
    
    search_dir = path if path and os.path.isdir(path) else "."
    
    files = []
    for f in os.listdir(search_dir):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            files.append(os.path.join(search_dir, f))
    
    return sorted(files)


def main():
    parser = argparse.ArgumentParser(
        description="Excel结构分析工具 - 自动分析工况Excel文件结构",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python excel_analyzer.py                              # 分析当前目录所有Excel
  python excel_analyzer.py data.xlsx                   # 分析指定文件
  python excel_analyzer.py data.xlsx --json           # 输出JSON格式
  python excel_analyzer.py data.xlsx -o report.txt    # 保存到文件
  python excel_analyzer.py data.xlsx --sample 5       # 每列显示5个样例
        """
    )
    parser.add_argument("path", nargs="?", default=None, help="Excel文件路径或目录")
    parser.add_argument("--json", "-j", action="store_true", help="输出JSON格式")
    parser.add_argument("--output", "-o", help="输出到文件")
    parser.add_argument("--sample", "-n", type=int, default=3, help="每列样例值数量(默认3)")
    
    args = parser.parse_args()
    
    analyzer = ExcelAnalyzer(sample_rows=args.sample)
    
    # 查找文件
    files = find_excel_files(args.path)
    
    if not files:
        print("未找到Excel文件")
        sys.exit(1)
    
    all_results = []
    
    for filepath in files:
        print(f"正在分析: {filepath}")
        try:
            result = analyzer.analyze_file(filepath)
            all_results.append(result)
        except Exception as e:
            print(f"  错误: {e}")
    
    # 输出结果
    if args.json:
        output = json.dumps(all_results, ensure_ascii=False, indent=2)
    else:
        outputs = []
        for result in all_results:
            outputs.append(analyzer.generate_report(result))
        output = "\n\n".join(outputs)
    
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f"\n报告已保存到: {args.output}")
    else:
        print("\n" + output)


if __name__ == "__main__":
    main()
